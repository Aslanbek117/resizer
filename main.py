import os

import image_slicer
from PIL import Image, ImageFilter
import urllib
import wget
import ssl
import pandas as pd
import openpyxl
from transliterate import translit, get_available_language_codes
import psycopg2
from psycopg2 import sql

ssl._create_default_https_context = ssl._create_unverified_context



def hyphen_split(a):
    if a.count("-") == 1:
        return a.split("-")[0]
    return "-".join(a.split("-", 2)[:2])

def hyphen_split_second(a):
    if a.count("/") == 1:
        return a.split("/")[0]
    return "/".join(a.split("/", 2))

# https://pixabay.com/images/download/kuala-lumpur-1820944.jpg?attachment

#https://pixabay.com/get/gca2b3f48f541307c43cb7d93d3eea1d06adeedaa0f335afc18efe57bbb41b2079dad39ac201668c683a0ea0e6642568f_1920.jpg?attachment=
def donwloadUrl(fileName, id):
    return "https://pixabay.com/images/download/"+fileName + "-"+id +".jpg?attachment"
#
# img = Image.open('./shadow_0.jpg', 'r').convert("RGB")
# img_w, img_h = img.size
# img2 = Image.open("./shadow_1.jpg", 'r').convert("RGB")
# offset2 = (700, 300)
#
# img3 = Image.open("./shadow_2.jpg", 'r').convert("RGB")
# offset3 = (1250, 300)
# background = Image.open("./background_3.jpg")
# bg_w, bg_h = background.size
# offset = (150,300)
# background.paste(img, offset)
# background.paste(img2, offset2)
# background.paste(img3, offset3)
# background.save('out_5.png')
#

# image_slicer.slice("./first.jpg", 4)
#
# im = Image.open('first.jpg').convert('RGB')
# im = im.crop((0,0, 600, 670))
# im.save('_0.png')

# im = Image.open('first.jpg').convert('RGB')
# im = im.crop((0,0, 600, 670))
# im.save('_0.png')
#
# im = Image.open('first.jpg').convert('RGB')
# im = im.crop((600,0, 1200, 670))
# im.save('_1.png')
#
# im = Image.open('first.jpg').convert('RGB')
# im = im.crop((1200,0, 1800, 670))
# im.save('_2.png')

# #
basewidth = 500
# img = Image.open('shadow_0.jpg').convert("RGB")
# wpercent = (basewidth/float(img.size[0]))
# hsize = int((float(img.size[1])*float(wpercent)))
# img = img.resize((basewidth,hsize), Image.ANTIALIAS)
# img.save('zaebal.jpg')
#
# basewidth = 500
# img = Image.open('_1.png').convert("RGB")
# wpercent = (basewidth/float(img.size[0]))
# hsize = int((float(img.size[1])*float(wpercent)))
# img = img.resize((basewidth,hsize), Image.ANTIALIAS)
# img.save('somepic_1.jpg')
#
# basewidth = 500
# img = Image.open('_2.png').convert("RGB")
# wpercent = (basewidth/float(img.size[0]))
# hsize = int((float(img.size[1])*float(wpercent)))
# img = img.resize((basewidth,hsize), Image.ANTIALIAS)
# img.save('somepic_2.jpg')
#image_slicer.slice('girl.png', 2)


def makeShadow(image, iterations, border, offset, backgroundColour, shadowColour):
    # image: base image to give a drop shadow
    # iterations: number of times to apply the blur filter to the shadow
    # border: border to give the image to leave space for the shadow
    # offset: offset of the shadow as [x,y]
    # backgroundCOlour: colour of the background
    # shadowColour: colour of the drop shadow

    # Calculate the size of the shadow's image
    fullWidth = image.size[0] + abs(offset[0]) + 2 * border
    fullHeight = image.size[1] + abs(offset[1]) + 2 * border

    # Create the shadow's image. Match the parent image's mode.
    shadow = Image.new(image.mode, (fullWidth, fullHeight), backgroundColour)

    # Place the shadow, with the required offset
    shadowLeft = border + max(offset[0], 0)  # if <0, push the rest of the image right
    shadowTop = border + max(offset[1], 0)  # if <0, push the rest of the image down
    # Paste in the constant colour
    shadow.paste(shadowColour,
                 [shadowLeft, shadowTop,
                  shadowLeft + image.size[0],
                  shadowTop + image.size[1]])

    # Apply the BLUR filter repeatedly
    for i in range(iterations):
        shadow = shadow.filter(ImageFilter.BLUR)

    # Paste the original image on top of the shadow
    imgLeft = border - min(offset[0], 0)  # if the shadow offset was <0, push right
    imgTop = border - min(offset[1], 0)  # if the shadow offset was <0, push down
    shadow.paste(image, (imgLeft, imgTop))
    shadow.save("somepic_0.jpg")
    return shadow

#
# img = Image.open('./somepic_0.jpg', 'r').convert("RGB")
#
# makeShadow(img, 5, 2, (0,0), "black", "white")

def dropShadow(fileName, image, offset=(5, 5), background=0xffffff, shadow=0x444444,
               border=8, iterations=10):
    """
    Add a gaussian blur drop shadow to an image.

    image       - The image to overlay on top of the shadow.
    offset      - Offset of the shadow from the image as an (x,y) tuple.  Can be
                  positive or negative.
    background  - Background colour behind the image.
    shadow      - Shadow colour (darkness).
    border      - Width of the border around the image.  This must be wide
                  enough to account for the blurring of the shadow.
    iterations  - Number of times to apply the filter.  More iterations
                  produce a more blurred shadow, but increase processing time.
    """

    # Create the backdrop image -- a box in the background colour with a
    # shadow on it.
    totalWidth = image.size[0] + abs(offset[0]) + 2 * border
    totalHeight = image.size[1] + abs(offset[1]) + 2 * border
    back = Image.new(image.mode, (totalWidth, totalHeight), background)

    # Place the shadow, taking into account the offset from the image
    shadowLeft = border + max(offset[0], 0)
    shadowTop = border + max(offset[1], 0)
    back.paste(shadow, [shadowLeft, shadowTop, shadowLeft + image.size[0],
                        shadowTop + image.size[1]])

    # Apply the filter to blur the edges of the shadow.  Since a small kernel
    # is used, the filter must be applied repeatedly to get a decent blur.
    n = 0
    while n < iterations:
        back = back.filter(ImageFilter.BLUR)
        n += 1

    # Paste the input image onto the shadow backdrop
    imageLeft = border - min(offset[0], 0)
    imageTop = border - min(offset[1], 0)
    back.paste(image, (imageLeft, imageTop))
    back.save(fileName)
    return back


# drop_shadow(Image.open('girl.png')).save('shadowed_boobs.png')

# dropShadow(Image.open("_1.png"))
# dropShadow(Image.open("somepic_2.jpg"), background=0xeeeeee, shadow=0x444444, offset=(0, 5))
# dropShadow(Image.open("_1.png"), background=0xeeeeee, shadow=0x444444, offset=(600, 0)).show()


def divide(offsetY, offsetX, fileIn, fileOut, width):
    im = Image.open(fileIn).convert('RGB')
    im = im.crop((offsetY,0, offsetX, width))
    im.save(fileOut)

def resize(_baseWidth, fileIn, fileOut):
    img = Image.open(fileIn).convert("RGB")
    wpercent = (basewidth / float(img.size[0]))
    hsize = int((float(img.size[1]) * float(wpercent)))
    img = img.resize((_baseWidth, hsize), Image.ANTIALIAS)
    w, h = img.size
    #img.save(fileOut)
    return h, img

def divide3(fileIn, fileOut, maxWidth, maxHeight):
    fixedOffset = maxWidth / 3
    baseShadowName = fileOut + "_"
    height = 0
    for i in range(3):
        out = fileOut + "_" + i.__str__() + ".jpg"
        resizeName = fileOut + "_" + i.__str__() + "_resize_500" + ".jpg"
        shadowName = fileOut + "_" + i.__str__() + "_resize_500_shadow" + ".jpg"
        if (i == 0):
            divide(0, fixedOffset, fileIn, fileOut + "_"+i.__str__() + ".jpg", maxHeight)
        elif (i == 1):
            divide(fixedOffset, fixedOffset*2, fileIn, fileOut + "_"+i.__str__() + ".jpg", maxHeight)
        elif (i == 2):
            divide(fixedOffset*2, maxWidth, fileIn, fileOut + "_"+i.__str__() + ".jpg", maxHeight)
        height, resizeImg = resize(500, out, resizeName)
        dropShadow(shadowName, resizeImg, background=0xeeeeee, shadow=0x444444, offset=(0, 5))

    img = Image.open(baseShadowName + "0_resize_500_shadow.jpg", 'r').convert("RGB")
    img2 = Image.open(baseShadowName + "1_resize_500_shadow.jpg", 'r').convert("RGB")
    img3 = Image.open(baseShadowName + "2_resize_500_shadow.jpg", 'r').convert("RGB")
    zaebal = int(1280 / 2  - height / 2)
    print(zaebal, height)
    offset2 = (int(fixedOffset) + 50, zaebal)
    offset3 = (int(fixedOffset) * 2 - 45, zaebal)
    background = Image.open("./background_3.jpg")
    offset = (150,zaebal)
    background.paste(img, offset)
    background.paste(img2, offset2)
    background.paste(img3, offset3)
    readyName = fileOut + "_complex_3.jpg"
    background.save(readyName)


def divide2(fileIn, fileOut, maxWidth, maxHeight):

    fixedOffset = maxWidth / 2
    baseShadowName = fileOut + "_"
    height = 0
    for i in range(2):
        out = fileOut + "_" + i.__str__() + ".jpg"
        resizeName = fileOut + "_" + i.__str__() + "_resize_500" + ".jpg"
        shadowName = fileOut + "_" + i.__str__() + "_resize_500_shadow" + ".jpg"
        if (i == 0):
            divide(0, fixedOffset, fileIn, fileOut + "_"+i.__str__() + ".jpg", maxHeight)
        elif (i == 1):
            divide(fixedOffset, fixedOffset*2, fileIn, fileOut + "_"+i.__str__() + ".jpg", maxHeight)
        height, resizeImg = resize(500, out, resizeName)
        dropShadow(shadowName, resizeImg, background=0xeeeeee, shadow=0x444444, offset=(0, 5))
    img = Image.open(baseShadowName + "0_resize_500_shadow.jpg", 'r').convert("RGB")
    img2 = Image.open(baseShadowName + "1_resize_500_shadow.jpg", 'r').convert("RGB")

    zaebal = int(1280 / 2  - height / 2)
    print(zaebal, height)
    offset2 = (int(fixedOffset) + 50, zaebal)

    background = Image.open("./background_3.jpg")
    offset = (75,zaebal)
    background.paste(img, offset)
    background.paste(img2, offset2)
    readyName = fileOut + "_complex_2.jpg"
    background.save(readyName)

def transform(fileIn, fileOut):
    im = Image.open(fileIn)

    # Target size parameters
    width = 1920
    height = 1080

    # Resize input image while keeping aspect ratio
    ratio = height / im.height
    im = im.resize((int(im.width * ratio), height))
    # Border parameters
    fill_color = (255, 255, 255)
    border_l = int((width - im.width) / 2)

    im_2 = Image.new('RGB', (width, height), fill_color)
    im_2.paste(im, (border_l, 0))
    im_2.save(fileOut +"_transform.jpg")

def divide2Low(fileIn, fileOut, maxWidth, maxHeight):
    fixedOffset = maxWidth / 2
    baseShadowName = fileOut + "_"
    height = 0
    for i in range(2):
        out = fileOut + "_" + i.__str__() + ".jpg"
        resizeName = fileOut + "_" + i.__str__() + "_resize_500" + ".jpg"
        shadowName = fileOut + "_" + i.__str__() + "_resize_500_shadow" + ".jpg"
        if (i == 0):
            divide(0, fixedOffset, fileIn, fileOut + "_"+i.__str__() + ".jpg", maxHeight)
        elif (i == 1):
            divide(fixedOffset, fixedOffset*2, fileIn, fileOut + "_"+i.__str__() + ".jpg", maxHeight)
        height, resizeImg = resize(int(maxWidth / 2) - 100, out, resizeName)
        dropShadow(shadowName, resizeImg, background=0xeeeeee, shadow=0x444444, offset=(0, 5))
    img = Image.open(baseShadowName + "0_resize_500_shadow.jpg", 'r').convert("RGB")
    img2 = Image.open(baseShadowName + "1_resize_500_shadow.jpg", 'r').convert("RGB")

    zaebal = int(1280 / 2  - height / 2)
    print(zaebal, height)
    offset2 = (int(fixedOffset) + 50, zaebal+100)

    background = Image.open("./background_3.jpg")
    offset = (75,zaebal)
    background.paste(img, offset)
    background.paste(img2, offset2)
    readyName = fileOut + "_complex_low_2.jpg"
    background.save(readyName)

def saveOriginal(fileIn, fileOut):
    im = Image.open(fileIn)
    im.save(fileOut+"_original.jpg")

def divide3Low(fileIn, fileOut, maxWidth, maxHeight):
    fixedOffset = maxWidth / 3
    baseShadowName = fileOut + "_"
    height = 0
    for i in range(3):
        out = fileOut + "_" + i.__str__() + ".jpg"
        resizeName = fileOut + "_" + i.__str__() + "_resize_500" + ".jpg"
        shadowName = fileOut + "_" + i.__str__() + "_resize_500_shadow" + ".jpg"
        if (i == 0):
            divide(0, fixedOffset, fileIn, fileOut + "_"+i.__str__() + ".jpg", maxHeight)
        elif (i == 1):
            divide(fixedOffset, fixedOffset*2, fileIn, fileOut + "_"+i.__str__() + ".jpg", maxHeight)
        elif (i == 2):
            divide(fixedOffset*2, maxWidth, fileIn, fileOut + "_"+i.__str__() + ".jpg", maxHeight)
        height, resizeImg = resize(500, out, resizeName)
        dropShadow(shadowName, resizeImg, background=0xeeeeee, shadow=0x444444, offset=(0, 5))

    img = Image.open(baseShadowName + "0_resize_500_shadow.jpg", 'r').convert("RGB")
    img2 = Image.open(baseShadowName + "1_resize_500_shadow.jpg", 'r').convert("RGB")
    img3 = Image.open(baseShadowName + "2_resize_500_shadow.jpg", 'r').convert("RGB")
    zaebal = int(1280 / 2  - height / 2)
    print(zaebal, height)
    offset2 = (int(fixedOffset) + 50, zaebal + 50)
    offset3 = (int(fixedOffset) * 2 - 45, zaebal + 100)
    background = Image.open("./background_3.jpg")
    offset = (150,zaebal)
    background.paste(img, offset)
    background.paste(img2, offset2)
    background.paste(img3, offset3)
    readyName = fileOut + "_complex_low_3.jpg"
    background.save(readyName)


# img = Image.open("london.jpg")
# w,h = img.size
# divide3("london.jpg", "f0", w, h)
# divide3Low("london.jpg", "f0", w, h)
# divide2("london.jpg", "f0", w, h)
# divide2Low("london.jpg", "lo", w,h)



# txt = "https://pixabay.com/photos/kuala-lumpur-petronas-twin-towers-1820944/"
# st = ""
# for i in txt:
#     if i.__str__().isdigit():
#         st+=i
#
# fileName = hyphen_split(txt).split("/")[-1]
# print(donwloadUrl(fileName, st))
# test


conn = psycopg2.connect(dbname='postgres', user='postgres',
                        password='1002574', host='localhost')

cursor = conn.cursor()
wrkbk = openpyxl.load_workbook("bridges.xlsx")
sh = wrkbk.active
# iterate through excel and display data
for i in range(1, sh.max_row + 1):
    # title = ""
    # fileName = ""
    # for j in range(1, sh.max_column + 1):
    #     cell_obj = sh.cell(row=i, column=j)
    #     # cell_obj.value
    #     if (j == 1):
    #         title = cell_obj.value
    #     else:
    #         fileName = cell_obj.value
    # tr = translit(title, "ru", reversed=True)
    # tr = tr.replace(" ", "_")
    # if not os.path.exists(tr):
    #     os.mkdir(tr)
    tr = "images/"
    title="images/"
    fileName ="bridge-53769_1920.jpg"
    if tr == "chlen":
        file = "./bridges/" + fileName
        img = Image.open(file)
        w, h = img.size
        divide3(file, "./" + tr + "/" + tr, w, h)
        divide3Low(file, "./" + tr + "/" + tr, w, h)
        divide2(file, "./" + tr + "/" + tr, w, h)
        divide2Low(file, "./" + tr + "/" + tr, w, h)
        transform(file, "./" + tr + "/" + tr)
        saveOriginal(file, "./" + tr + "/" + tr)
        with conn.cursor() as cursor:
            zaebal = "./" + tr + "/" + tr
            conn.autocommit = True
            values = [
                (file, title, 'Мосты', tr, zaebal + "_0.jpg", zaebal + "_1.jpg", zaebal + "_2.jpg",
                 zaebal + "_0_resize_500_shadow.jpg", zaebal + "_1_resize_500_shadow.jpg",
                 zaebal + "_2_resize_500_shadow.jpg", zaebal + "_complex_2.jpg", zaebal + "_complex_low_2.jpg",
                 zaebal + "_complex_3.jpg", zaebal + "_complex_low_3.jpg", zaebal + "_transform.jpg",
                 zaebal + "_original.jpg")
            ]
            insert = sql.SQL(
                'INSERT INTO pictures.catalog (path, title, category, directory_name, resize_0, resize_1, resize_2, resize_0_shadow, resize_1_shadow, resize_2_shadow, complex_2, complex_2_low, complex_3, complex_3_low, transform, original) VALUES {}').format(
                sql.SQL(',').join(map(sql.Literal, values))
            )
            cursor.execute(insert)

